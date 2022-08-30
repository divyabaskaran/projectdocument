import openpyxl


def readData(file, sheetname, rownum, colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    inputvalue = sheet.cell(row=rownum, column=colnum).value
    return inputvalue


def getconfig(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet['A2'].value

def getdriverconfig(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet['B2'].value

def getlogindata(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]

    return sheet['A2'].value, sheet['B2'].value

def gettestdatadoc(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]

    return sheet['C2'].value, sheet['D2'].value

def gettestdataimpfol(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]

    return sheet['E2'].value, sheet['F2'].value

def gettabname(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet['B2'].value


def writedata(file, sheetname, row, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]

    sheet['C' + str(row)] = data
    workbook.save(file)

def foldercheck(file, sheetname, row, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]

    sheet['D' + str(row)] = data
    workbook.save(file)


def writeerror(file, sheetname, row, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]

    sheet['D' + str(row)] = data
    workbook.save(file)



def removedata(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    for i in range(2, 11):
        sheet['C' + str(i)] = ''
    for j in range(2, 11):
        sheet['D' + str(j)] = ''

    workbook.save(file)

def projectno(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet['A2'].value